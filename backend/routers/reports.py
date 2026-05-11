from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, date, timedelta
from typing import Optional
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import models
from database import get_db

router = APIRouter(prefix="/relatorios", tags=["Relatórios"])


@router.get("/resumo")
def resumo(db: Session = Depends(get_db)):
    total       = db.query(func.count(models.Ferramenta.id)).scalar() or 0
    disponiveis = db.query(func.count(models.Ferramenta.id)).filter(models.Ferramenta.estado == "disponivel").scalar() or 0
    emprestadas = db.query(func.count(models.Ferramenta.id)).filter(models.Ferramenta.estado == "emprestada").scalar() or 0
    manutencao  = db.query(func.count(models.Ferramenta.id)).filter(models.Ferramenta.estado == "manutencao").scalar() or 0
    emp_abertos = db.query(func.count(models.Emprestimo.id)).filter(models.Emprestimo.devolvida == False).scalar() or 0

    por_categoria = (
        db.query(models.Categoria.nome, func.count(models.Ferramenta.id))
        .outerjoin(models.Ferramenta, models.Ferramenta.categoria_id == models.Categoria.id)
        .group_by(models.Categoria.id)
        .all()
    )
    # Extintores resumo
    hoje = date.today()
    ext_vencidos  = db.query(func.count(models.Extintor.id)).filter(
        models.Extintor.ativo == True,
        models.Extintor.data_validade < hoje,
    ).scalar() or 0
    ext_vence_30d = db.query(func.count(models.Extintor.id)).filter(
        models.Extintor.ativo == True,
        models.Extintor.data_validade >= hoje,
        models.Extintor.data_validade <= hoje + timedelta(days=30),
    ).scalar() or 0

    return {
        "total_ferramentas": total,
        "disponiveis":       disponiveis,
        "emprestadas":       emprestadas,
        "em_manutencao":     manutencao,
        "emprestimos_abertos": emp_abertos,
        "extintores_vencidos":  ext_vencidos,
        "extintores_vence_30d": ext_vence_30d,
        "por_categoria": [{"categoria": n, "total": q} for n, q in por_categoria],
    }


@router.get("/ferramentas-por-igreja")
def ferramentas_por_igreja(db: Session = Depends(get_db)):
    rows = (
        db.query(models.Igreja.nome, models.Igreja.id, func.count(models.Ferramenta.id).label("total"))
        .outerjoin(models.Ferramenta, models.Ferramenta.igreja_id == models.Igreja.id)
        .group_by(models.Igreja.id)
        .order_by(func.count(models.Ferramenta.id).desc())
        .all()
    )
    sem = db.query(func.count(models.Ferramenta.id)).filter(models.Ferramenta.igreja_id == None).scalar() or 0
    resultado = [{"igreja": n, "id": i, "total": t} for n, i, t in rows]
    if sem > 0:
        resultado.append({"igreja": "Sem igreja", "id": None, "total": sem})
    return resultado

@router.get("/uso-por-setor")
def uso_por_setor(db: Session = Depends(get_db)):
    dados = (
        db.query(
            models.Usuario.setor,
            func.count(models.Emprestimo.id).label("total")
        )
        .join(
            models.Usuario,
            models.Usuario.id == models.Emprestimo.responsavel_id,
            isouter=True
        )
        .group_by(models.Usuario.setor)
        .all()
    )

    return [
        {
            "setor": d[0] or "Sem setor",
            "total": d[1]
        }
        for d in dados
    ]
    
@router.get("/solicitacoes-pendentes")
def solicitacoes_pendentes(db: Session = Depends(get_db)):
    solicitacoes = (
        db.query(models.Solicitacao)
        .filter(models.Solicitacao.status == "pendente")
        .all()
    )

    return [
        {
            "id": s.id,
            "item_nome": s.item_nome,
            "setor": s.setor,
            "tipo": s.tipo,
            "status": s.status,
            "criado_em": s.criado_em
        }
        for s in solicitacoes
    ]

@router.get("/reservas-ativas")
def reservas_ativas(db: Session = Depends(get_db)):
    reservas = (
        db.query(models.Reserva)
        .filter(models.Reserva.status == "ativa")
        .all()
    )

    return [
        {
            "id": r.id,
            "ferramenta_id": r.ferramenta_id,
            "usuario_id": r.usuario_id,
            "data_inicio": r.data_inicio,
            "data_fim": r.data_fim,
            "status": r.status
        }
        for r in reservas
    ]
    
@router.get("/extintores-vencidos")
def extintores_vencidos(
    igreja_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    hoje = date.today()
    q = db.query(models.Extintor).filter(
        models.Extintor.ativo == True,
        models.Extintor.data_validade != None,
        models.Extintor.data_validade < hoje,
    )
    if igreja_id:
        q = q.filter(models.Extintor.igreja_id == igreja_id)
    return [
        {
            "id":           e.id,
            "patrimonio":   e.patrimonio,
            "tipo":         e.tipo,
            "capacidade":   e.capacidade,
            "localizacao":  e.localizacao,
            "data_validade":e.data_validade.isoformat(),
            "dias_vencido": (hoje - e.data_validade).days,
            "igreja":       e.igreja.nome if e.igreja else "Sem igreja",
        }
        for e in q.order_by(models.Extintor.data_validade).all()
    ]


@router.get("/extintores-por-igreja")
def extintores_por_igreja(db: Session = Depends(get_db)):
    hoje = date.today()
    igrejas = db.query(models.Igreja).filter(models.Igreja.ativo == True).order_by(models.Igreja.nome).all()
    resultado = []
    for ig in igrejas:
        exts = db.query(models.Extintor).filter(
            models.Extintor.igreja_id == ig.id, models.Extintor.ativo == True
        ).all()
        vencidos  = sum(1 for e in exts if e.data_validade and e.data_validade < hoje)
        vence_30d = sum(1 for e in exts if e.data_validade and hoje <= e.data_validade <= hoje + timedelta(days=30))
        ok        = sum(1 for e in exts if e.data_validade and e.data_validade > hoje + timedelta(days=30))
        resultado.append({
            "igreja": ig.nome, "id": ig.id,
            "total": len(exts), "vencidos": vencidos, "vence_30d": vence_30d, "ok": ok,
        })
    return resultado


@router.get("/mais-usadas")
def mais_usadas(
    categoria_id: Optional[int]      = Query(None),
    igreja_id:    Optional[int]      = Query(None),
    limit: int = 10,
    db: Session = Depends(get_db),
):
    q = (
        db.query(models.Ferramenta.nome, models.Ferramenta.id, func.count(models.Emprestimo.id).label("usos"))
        .join(models.Emprestimo, models.Emprestimo.ferramenta_id == models.Ferramenta.id)
    )
    if categoria_id: q = q.filter(models.Ferramenta.categoria_id == categoria_id)
    if igreja_id:    q = q.filter(models.Ferramenta.igreja_id == igreja_id)
    rows = q.group_by(models.Ferramenta.id).order_by(func.count(models.Emprestimo.id).desc()).limit(limit).all()
    return [{"ferramenta": n, "id": i, "usos": u} for n, i, u in rows]


@router.get("/atrasados")
def atrasados(db: Session = Depends(get_db)):
    hoje = datetime.utcnow()
    emp = (
        db.query(models.Emprestimo)
        .filter(
            models.Emprestimo.devolvida == False,
            models.Emprestimo.data_retorno_prevista != None,
            models.Emprestimo.data_retorno_prevista < hoje,
        )
        .all()
    )
    return [
        {
            "id":                    e.id,
            "ferramenta":            e.ferramenta.nome if e.ferramenta else "",
            "igreja":                e.ferramenta.igreja.nome if e.ferramenta and e.ferramenta.igreja else None,
            "responsavel":           e.responsavel.nome if e.responsavel else "",
            "data_saida":            e.data_saida.isoformat(),
            "data_retorno_prevista": e.data_retorno_prevista.isoformat(),
            "dias_atraso":           (hoje - e.data_retorno_prevista).days,
        }
        for e in emp
    ]


@router.get("/exportar/ferramentas")
def exportar_ferramentas(
    categoria_id: Optional[int] = Query(None),
    estado:       Optional[str] = Query(None),
    igreja_id:    Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(models.Ferramenta)
    if categoria_id: q = q.filter(models.Ferramenta.categoria_id == categoria_id)
    if estado:       q = q.filter(models.Ferramenta.estado == estado)
    if igreja_id:    q = q.filter(models.Ferramenta.igreja_id == igreja_id)
    ferramentas = q.order_by(models.Ferramenta.nome).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ferramentas"
    hfill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True)
    headers = ["ID","Nome","Descrição","Nº Série","Localização","Estado","Categoria","Igreja","Cadastrado em"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    estado_map = {"disponivel":"Disponível","emprestada":"Emprestada","manutencao":"Manutenção"}
    for row, f in enumerate(ferramentas, 2):
        ws.cell(row=row, column=1, value=f.id)
        ws.cell(row=row, column=2, value=f.nome)
        ws.cell(row=row, column=3, value=f.descricao or "")
        ws.cell(row=row, column=4, value=f.numero_serie or "")
        ws.cell(row=row, column=5, value=f.localizacao or "")
        ws.cell(row=row, column=6, value=estado_map.get(f.estado, f.estado))
        ws.cell(row=row, column=7, value=f.categoria.nome if f.categoria else "")
        ws.cell(row=row, column=8, value=f.igreja.nome   if f.igreja    else "")
        ws.cell(row=row, column=9, value=f.criado_em.strftime("%d/%m/%Y"))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 4, 40
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=ferramentas_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )


@router.get("/exportar/emprestimos")
def exportar_emprestimos(
    devolvida: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(models.Emprestimo)
    if devolvida is not None:
        q = q.filter(models.Emprestimo.devolvida == devolvida)
    emprestimos = q.order_by(models.Emprestimo.data_saida.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empréstimos"
    hfill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True)
    headers = ["ID","Ferramenta","Igreja","Responsável","Saída","Retorno prev.","Retorno real","Status","Observações"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    for row, e in enumerate(emprestimos, 2):
        ws.cell(row=row, column=1, value=e.id)
        ws.cell(row=row, column=2, value=e.ferramenta.nome if e.ferramenta else "")
        ws.cell(row=row, column=3, value=e.ferramenta.igreja.nome if e.ferramenta and e.ferramenta.igreja else "")
        ws.cell(row=row, column=4, value=e.responsavel.nome if e.responsavel else "")
        ws.cell(row=row, column=5, value=e.data_saida.strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row, column=6, value=e.data_retorno_prevista.strftime("%d/%m/%Y") if e.data_retorno_prevista else "")
        ws.cell(row=row, column=7, value=e.data_retorno_real.strftime("%d/%m/%Y %H:%M") if e.data_retorno_real else "")
        ws.cell(row=row, column=8, value="Devolvida" if e.devolvida else "Em aberto")
        ws.cell(row=row, column=9, value=e.observacoes or "")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 4, 40
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=emprestimos_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )