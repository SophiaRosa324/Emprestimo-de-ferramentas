from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
from datetime import date, timedelta
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi.responses import StreamingResponse

import models
import schemas
from database import get_db

router = APIRouter(prefix="/extintores", tags=["Extintores"])


# ── Serialização (inclui campos calculados) ────────────────
def _s(e: models.Extintor) -> dict:
    hoje = date.today()
    sv = e.status_validade
    dpv = e.dias_para_vencer
    return {
        "id":               e.id,
        "patrimonio":       e.patrimonio,
        "tipo":             e.tipo,
        "capacidade":       e.capacidade,
        "localizacao":      e.localizacao,
        "data_validade":    e.data_validade.isoformat()     if e.data_validade     else None,
        "data_ultima_carga":e.data_ultima_carga.isoformat() if e.data_ultima_carga else None,
        "observacoes":      e.observacoes,
        "ativo":            e.ativo,
        "igreja_id":        e.igreja_id,
        "igreja_nome":      e.igreja.nome if e.igreja else None,
        "status_validade":  sv,
        "dias_para_vencer": dpv,
        "criado_em":        e.criado_em.isoformat(),
        "atualizado_em":    e.atualizado_em.isoformat() if e.atualizado_em else None,
    }


# ── Rotas ──────────────────────────────────────────────────
@router.get("/dashboard/resumo")
def dashboard_extintores(
    igreja_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(models.Extintor).filter(models.Extintor.ativo == True)
    if igreja_id:
        q = q.filter(models.Extintor.igreja_id == igreja_id)
    todos = q.all()

    hoje = date.today()
    vencidos  = [e for e in todos if e.data_validade and e.data_validade < hoje]
    vence_30d = [e for e in todos if e.data_validade and hoje <= e.data_validade <= hoje + timedelta(days=30)]
    ok        = [e for e in todos if e.data_validade and e.data_validade > hoje + timedelta(days=30)]
    sem_val   = [e for e in todos if not e.data_validade]

    por_tipo = {}
    for e in todos:
        por_tipo[e.tipo] = por_tipo.get(e.tipo, 0) + 1

    por_igreja = {}
    for e in todos:
        nome = e.igreja.nome if e.igreja else "Sem igreja"
        por_igreja[nome] = por_igreja.get(nome, 0) + 1

    return {
        "total":        len(todos),
        "vencidos":     len(vencidos),
        "vence_30d":    len(vence_30d),
        "ok":           len(ok),
        "sem_validade": len(sem_val),
        "por_tipo":     [{"tipo": t, "total": c} for t, c in sorted(por_tipo.items(), key=lambda x: -x[1])],
        "por_igreja":   [{"igreja": i, "total": c} for i, c in sorted(por_igreja.items(), key=lambda x: -x[1])],
        "proximos_vencer": [_s(e) for e in sorted(vence_30d, key=lambda x: x.data_validade)],
        "vencidos_lista":  [_s(e) for e in sorted(vencidos,  key=lambda x: x.data_validade)],
    }


@router.get("/vencendo")
def listar_vencendo(
    dias: int          = Query(30),
    igreja_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    hoje = date.today()
    limite = hoje + timedelta(days=dias)
    q = db.query(models.Extintor).filter(
        models.Extintor.ativo == True,
        models.Extintor.data_validade != None,
        models.Extintor.data_validade <= limite,
    )
    if igreja_id:
        q = q.filter(models.Extintor.igreja_id == igreja_id)
    return [_s(e) for e in q.order_by(models.Extintor.data_validade).all()]


@router.get("/")
def listar_extintores(
    igreja_id:    Optional[int]  = Query(None),
    status:       Optional[str]  = Query(None),
    tipo:         Optional[str]  = Query(None),
    ativo:        Optional[bool] = Query(True),
    busca:        Optional[str]  = Query(None),
    skip:  int = 0,
    limit: int = 200,
    db: Session = Depends(get_db),
):
    q = db.query(models.Extintor)
    if ativo is not None:
        q = q.filter(models.Extintor.ativo == ativo)
    if igreja_id:
        q = q.filter(models.Extintor.igreja_id == igreja_id)
    if tipo:
        q = q.filter(models.Extintor.tipo.ilike(f"%{tipo}%"))
    if busca:
        q = q.filter(
            models.Extintor.patrimonio.ilike(f"%{busca}%") |
            models.Extintor.localizacao.ilike(f"%{busca}%") |
            models.Extintor.tipo.ilike(f"%{busca}%")
        )
    todos = q.order_by(models.Extintor.data_validade.asc().nullslast()).offset(skip).limit(limit).all()
    resultado = [_s(e) for e in todos]
    if status:
        resultado = [e for e in resultado if e["status_validade"] == status]
    return resultado


@router.get("/{id}")
def obter_extintor(id: int, db: Session = Depends(get_db)):
    e = db.query(models.Extintor).filter(models.Extintor.id == id).first()
    if not e:
        raise HTTPException(status_code=404, detail="Extintor não encontrado.")
    return _s(e)


@router.post("/", status_code=201)
def criar_extintor(dados: schemas.ExtintorCreate, db: Session = Depends(get_db)):
    dup = db.query(models.Extintor).filter(
        models.Extintor.patrimonio == dados.patrimonio.strip()
    ).first()
    if dup:
        raise HTTPException(status_code=400, detail=f"Patrimônio '{dados.patrimonio}' já cadastrado.")
    e = models.Extintor(**dados.model_dump())
    e.patrimonio = dados.patrimonio.strip()
    db.add(e)
    db.commit()
    db.refresh(e)
    return _s(e)


@router.put("/{id}")
def atualizar_extintor(id: int, dados: schemas.ExtintorUpdate, db: Session = Depends(get_db)):
    e = db.query(models.Extintor).filter(models.Extintor.id == id).first()
    if not e:
        raise HTTPException(status_code=404, detail="Extintor não encontrado.")
    if dados.patrimonio and dados.patrimonio.strip() != e.patrimonio:
        dup = db.query(models.Extintor).filter(
            models.Extintor.patrimonio == dados.patrimonio.strip(),
            models.Extintor.id != id,
        ).first()
        if dup:
            raise HTTPException(status_code=400, detail=f"Patrimônio '{dados.patrimonio}' já existe.")
    for campo, valor in dados.model_dump(exclude_unset=True).items():
        setattr(e, campo, valor)
    db.commit()
    db.refresh(e)
    return _s(e)


@router.delete("/{id}", status_code=204)
def deletar_extintor(id: int, db: Session = Depends(get_db)):
    e = db.query(models.Extintor).filter(models.Extintor.id == id).first()
    if not e:
        raise HTTPException(status_code=404, detail="Extintor não encontrado.")
    db.delete(e)
    db.commit()


@router.get("/exportar/excel")
def exportar_extintores(
    igreja_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    hoje = date.today()
    q = db.query(models.Extintor).filter(models.Extintor.ativo == True)
    if igreja_id:
        q = q.filter(models.Extintor.igreja_id == igreja_id)
    todos = q.order_by(models.Extintor.data_validade.asc().nullslast()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extintores"

    red_fill    = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    yellow_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    hfill       = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    hfont       = Font(color="FFFFFF", bold=True)

    headers = ["Patrimônio","Tipo","Capacidade","Localização","Igreja","Validade","Última carga","Status","Dias p/ vencer"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    status_label = {"vencido":"VENCIDO","vence_30d":"Vence em 30d","ok":"Em dia","sem_validade":"Sem validade"}

    for row, e in enumerate(todos, 2):
        sv  = e.status_validade
        dpv = e.dias_para_vencer
        ws.cell(row=row, column=1, value=e.patrimonio)
        ws.cell(row=row, column=2, value=e.tipo)
        ws.cell(row=row, column=3, value=e.capacidade or "")
        ws.cell(row=row, column=4, value=e.localizacao or "")
        ws.cell(row=row, column=5, value=e.igreja.nome if e.igreja else "")
        ws.cell(row=row, column=6, value=e.data_validade.strftime("%d/%m/%Y")     if e.data_validade     else "")
        ws.cell(row=row, column=7, value=e.data_ultima_carga.strftime("%d/%m/%Y") if e.data_ultima_carga else "")
        ws.cell(row=row, column=8, value=status_label.get(sv, sv))
        ws.cell(row=row, column=9, value=dpv)
        fill = red_fill if sv == "vencido" else yellow_fill if sv == "vence_30d" else None
        if fill:
            for c in range(1, 10):
                ws.cell(row=row, column=c).fill = fill

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 4, 35
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import datetime
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=extintores_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )